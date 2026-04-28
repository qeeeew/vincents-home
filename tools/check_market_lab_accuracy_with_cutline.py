import csv
import re
from pathlib import Path

from openpyxl import load_workbook

import generate_market_lab_pdf_test_cases as estimator


ROOT = Path("/Users/iyulha/Desktop/Vincent's Home")
CUTLINE_XLSX = Path("/Users/iyulha/Desktop/sh 청년매입임대주택 자료/2024 1차 커트라인.xlsx")
OUTPUT_PATH = ROOT / "exports" / "market_lab_2024_1_accuracy_check.csv"

RANK_LABEL = {
    1: "1순위",
    2: "2순위",
    3: "3순위",
    "1": "1순위",
    "2": "2순위",
    "3": "3순위",
    "-": None,
    None: None,
}


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace(" ", "")
    text = text.replace(",", "")
    text = text.replace("외1필지", "외1필지")
    text = text.replace("외2필지", "외2필지")
    text = text.replace("외3필지", "외3필지")
    text = re.sub(r"\s+", "", text)
    return text


def normalize_housing_name(value):
    text = str(value or "").strip()
    match = re.match(r"^(.*?)\[(.*?)\]$", text)
    if not match:
        return text, "", ""
    base_name, bracket = match.groups()
    gender = ""
    housing_type = bracket
    if "/" in bracket:
        left, right = bracket.split("/", 1)
        if left in ("남성", "여성"):
            gender = left
            housing_type = right
    elif bracket in ("남성", "여성", "-"):
        gender = "" if bracket == "-" else bracket
        housing_type = ""
    elif bracket == "-":
        housing_type = ""
    return base_name.strip(), housing_type.strip(), gender.strip()


def load_cutline_rows():
    wb = load_workbook(CUTLINE_XLSX, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for row_no, row in enumerate(ws.iter_rows(min_row=10, values_only=True), start=10):
        if not any(row):
            continue
        source_kind = row[1]
        district = row[2]
        housing_name_raw = row[3]
        address = row[4]
        gender_cell = row[5]
        winning_rank = RANK_LABEL.get(row[6], None)
        winning_score = row[7] if isinstance(row[7], (int, float)) else None
        reserve_rank = RANK_LABEL.get(row[8], None)
        reserve_score = row[9] if isinstance(row[9], (int, float)) else None
        if not source_kind or not district or not address:
            continue
        base_name, housing_type, gender_from_name = normalize_housing_name(housing_name_raw)
        gender = "" if gender_cell in (None, "-") else str(gender_cell).strip()
        if not gender:
            gender = gender_from_name
        rows.append(
            {
                "xlsx_row_no": row_no,
                "kind": str(source_kind).strip(),
                "district": str(district).strip(),
                "housingName": base_name,
                "housingType": housing_type,
                "gender": gender,
                "address": str(address).strip(),
                "winningRank": winning_rank,
                "winningScore": winning_score,
                "reserveRank": reserve_rank,
                "reserveScore": reserve_score,
            }
        )
    return rows


def build_pdf_index():
    index = {}
    for row in estimator.parse_pdf_rows():
        key = (
            normalize_text(row["kind"]),
            normalize_text(row["district"]),
            normalize_text(row["address"]),
        )
        index.setdefault(key, []).append(row)
    return index


def choose_pdf_match(cutline_row, candidates):
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    target_gender = normalize_text(cutline_row["gender"])
    target_type = normalize_text(cutline_row["housingType"])
    target_name = normalize_text(cutline_row["housingName"])

    for candidate in candidates:
        if normalize_text(candidate["gender"]) == target_gender and normalize_text(candidate["housingType"]) == target_type:
            return candidate
    for candidate in candidates:
        if normalize_text(candidate["housingType"]) == target_type and normalize_text(candidate["housingName"]) == target_name:
            return candidate
    for candidate in candidates:
        if normalize_text(candidate["housingName"]) == target_name:
            return candidate
    return candidates[0]


def classify_accuracy(probability, threshold):
    return "통과" if probability >= threshold else "미달"


def main():
    estimator.BOOTSTRAP_RUNS = 80
    market_lab_data = estimator.load_market_lab_data()
    pdf_index = build_pdf_index()
    cutline_rows = load_cutline_rows()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "xlsx_row_no",
        "matched",
        "source_kind",
        "source_district",
        "source_housing_name",
        "source_housing_type",
        "source_gender",
        "source_supply",
        "source_applicants",
        "source_competition_ratio",
        "actual_winning_rank",
        "actual_winning_score",
        "pred_win_rate_at_winning_cut",
        "pred_reserve_rate_at_winning_cut",
        "winning_cut_check_50",
        "winning_cut_check_70",
        "actual_reserve_rank",
        "actual_reserve_score",
        "pred_win_rate_at_reserve_cut",
        "pred_reserve_rate_at_reserve_cut",
        "reserve_cut_check_50",
        "reserve_cut_check_70",
        "source_address",
    ]

    matched_count = 0
    with OUTPUT_PATH.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()

        for row in cutline_rows:
            key = (
                normalize_text(row["kind"]),
                normalize_text(row["district"]),
                normalize_text(row["address"]),
            )
            pdf_row = choose_pdf_match(row, pdf_index.get(key, []))
            if pdf_row:
                matched_count += 1

            winning_result = None
            reserve_result = None

            if pdf_row and row["winningRank"] and row["winningScore"] is not None:
                winning_result = estimator.calculate_estimate(
                    {
                        "rank": row["winningRank"],
                        "score": float(row["winningScore"]),
                        "applicants": int(pdf_row["applicants"]),
                        "supply": int(pdf_row["supply"]),
                        "includeFakeSupport": True,
                    },
                    market_lab_data,
                    300000 + row["xlsx_row_no"],
                )

            if pdf_row and row["reserveRank"] and row["reserveScore"] is not None:
                reserve_result = estimator.calculate_estimate(
                    {
                        "rank": row["reserveRank"],
                        "score": float(row["reserveScore"]),
                        "applicants": int(pdf_row["applicants"]),
                        "supply": int(pdf_row["supply"]),
                        "includeFakeSupport": True,
                    },
                    market_lab_data,
                    400000 + row["xlsx_row_no"],
                )

            writer.writerow(
                {
                    "xlsx_row_no": row["xlsx_row_no"],
                    "matched": "Y" if pdf_row else "N",
                    "source_kind": row["kind"],
                    "source_district": row["district"],
                    "source_housing_name": row["housingName"],
                    "source_housing_type": row["housingType"],
                    "source_gender": row["gender"],
                    "source_supply": pdf_row["supply"] if pdf_row else "",
                    "source_applicants": pdf_row["applicants"] if pdf_row else "",
                    "source_competition_ratio": pdf_row["competitionRatio"] if pdf_row else "",
                    "actual_winning_rank": row["winningRank"] or "",
                    "actual_winning_score": row["winningScore"] if row["winningScore"] is not None else "",
                    "pred_win_rate_at_winning_cut": round(winning_result["winRate"], 6) if winning_result else "",
                    "pred_reserve_rate_at_winning_cut": round(winning_result["reserveRate"], 6) if winning_result else "",
                    "winning_cut_check_50": classify_accuracy(winning_result["winRate"], 0.5) if winning_result else "",
                    "winning_cut_check_70": classify_accuracy(winning_result["winRate"], 0.7) if winning_result else "",
                    "actual_reserve_rank": row["reserveRank"] or "",
                    "actual_reserve_score": row["reserveScore"] if row["reserveScore"] is not None else "",
                    "pred_win_rate_at_reserve_cut": round(reserve_result["winRate"], 6) if reserve_result else "",
                    "pred_reserve_rate_at_reserve_cut": round(reserve_result["reserveRate"], 6) if reserve_result else "",
                    "reserve_cut_check_50": classify_accuracy(reserve_result["reserveRate"], 0.5) if reserve_result else "",
                    "reserve_cut_check_70": classify_accuracy(reserve_result["reserveRate"], 0.7) if reserve_result else "",
                    "source_address": row["address"],
                }
            )

    print(f"cutline_rows={len(cutline_rows)}")
    print(f"matched_rows={matched_count}")
    print(f"output={OUTPUT_PATH}")


if __name__ == "__main__":
    main()
