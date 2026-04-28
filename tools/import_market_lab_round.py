from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl
from pypdf import PdfReader


ROOT = Path("/Users/iyulha/Desktop/Vincent's Home")
DATA_FILE = ROOT / "market-lab-data.js"

GENDER_VALUES = {"남성", "여성"}
SCORE_RE = re.compile(r"(\d+)")


def round_half_up(value: float, digits: int = 1) -> float:
    quant = "0." + ("0" * (digits - 1)) + "1"
    return float(Decimal(str(value)).quantize(Decimal(quant), rounding=ROUND_HALF_UP))


def load_records() -> list[dict]:
    text = DATA_FILE.read_text(encoding="utf-8").strip()
    payload = text.split("=", 1)[1].strip()
    if payload.endswith(";"):
        payload = payload[:-1]
    return json.loads(payload)


def save_records(records: list[dict]) -> None:
    cleaned_records = [
        {key: value for key, value in record.items() if not key.startswith("_")}
        for record in records
    ]
    serialized = json.dumps(cleaned_records, ensure_ascii=False, separators=(",", ":"))
    DATA_FILE.write_text(f"window.marketLabData = {serialized};\n", encoding="utf-8")


def clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "-":
        return None
    return text


def parse_score(value) -> int | None:
    text = clean_text(value)
    if not text:
        return None
    match = SCORE_RE.search(text)
    return int(match.group(1)) if match else None


def parse_rank(value) -> str | None:
    text = clean_text(value)
    return text


def parse_name_descriptor(cell_value: str) -> tuple[str, str | None, str | None]:
    text = str(cell_value).strip()
    match = re.match(r"^(?P<name>.+?)\[(?P<descriptor>.*)\]$", text)
    if not match:
        return text, None, None

    name = match.group("name").strip()
    raw_descriptor = match.group("descriptor").strip()
    if raw_descriptor in {"", "-"}:
        return name, None, None

    parts = [part.strip() for part in re.split(r"[,/]", raw_descriptor) if part.strip()]
    genders = [part for part in parts if part in GENDER_VALUES]
    non_genders = [part for part in parts if part not in GENDER_VALUES]

    if genders and non_genders:
        gender = genders[-1]
        return name, f"{''.join(non_genders)}({gender})", gender
    if genders:
        gender = genders[-1]
        return name, gender, gender
    return name, raw_descriptor, None


def normalize_descriptor(housing_type: str | None, gender: str | None) -> str:
    if gender:
        if not housing_type or housing_type == gender:
            return gender
        if gender in housing_type:
            return housing_type
        return f"{housing_type}({gender})"
    return housing_type or "-"


def normalize_address(address: str) -> str:
    return re.sub(r"\s+", "", address)


def split_type_gender(type_text: str | None, gender_text: str | None) -> tuple[str | None, str | None]:
    raw_type = type_text.strip() if type_text else None
    raw_gender = gender_text.strip() if gender_text else None

    if raw_type and " " in raw_type and raw_gender is None:
        left, right = raw_type.rsplit(" ", 1)
        if right in {"-", "남성", "여성"}:
            raw_type = left.strip() or None
            raw_gender = right

    gender = None if raw_gender in {None, "-"} else raw_gender
    housing_type = None if raw_type in {None, "-"} else raw_type
    return housing_type, gender


def parse_xlsx_records(xlsx_path: Path, round_key: str, round_label: str, notice_date: str) -> list[dict]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    records: list[dict] = []
    order = 0
    for row in sheet.iter_rows(min_row=9, values_only=True):
        kind = clean_text(row[1])
        district = clean_text(row[2])
        name_cell = clean_text(row[3])
        address = clean_text(row[4])
        winning_rank = parse_rank(row[5])
        winning_score = parse_score(row[6])
        reserve_rank = parse_rank(row[7])
        reserve_score = parse_score(row[8])

        if not any([kind, district, name_cell, address]):
            continue

        if not all([kind, district, name_cell, address, winning_rank]) or winning_score is None:
            raise ValueError(f"Incomplete xlsx row: {row}")

        housing_name, housing_type, gender = parse_name_descriptor(name_cell)
        if reserve_rank == "-":
            reserve_rank = None

        records.append(
            {
                "_order": order,
                "roundKey": round_key,
                "roundLabel": round_label,
                "noticeDate": notice_date,
                "kind": kind,
                "district": district,
                "housingName": housing_name,
                "housingType": housing_type,
                "gender": gender,
                "address": address,
                "winningRank": winning_rank,
                "winningScore": winning_score,
                "reserveRank": reserve_rank,
                "reserveScore": reserve_score,
            }
        )
        order += 1

    return records


def group_page_fragments(page) -> list[tuple[float, list[tuple[float, str]]]]:
    fragments: list[tuple[float, float, str]] = []

    def visitor(text, cm, tm, font_dict, font_size):
        value = text.strip()
        if not value:
            return
        x = round(tm[4], 1)
        y = round(tm[5], 1)
        if y <= 0:
            return
        fragments.append((y, x, value))

    page.extract_text(visitor_text=visitor)
    rows_by_y: dict[float, list[tuple[float, str]]] = {}
    for y, x, value in fragments:
        rows_by_y.setdefault(y, []).append((x, value))

    grouped = []
    for y, row in rows_by_y.items():
        grouped.append((y, sorted(row, key=lambda item: item[0])))
    return sorted(grouped, key=lambda item: item[0])


def parse_rank_line(text: str) -> tuple[str, int | None] | None:
    subtotal = re.match(r"^소계\b", text)
    if subtotal:
        return "소계", None

    match = re.match(r"^(1순위|2순위|3순위)\s+(\d+)", text)
    if not match:
        return None
    return match.group(1), int(match.group(2))


def build_group_key(kind: str, district: str, housing_name: str, housing_type: str | None, gender: str | None, address: str) -> tuple[str, str, str, str, str]:
    return (
        kind,
        district,
        housing_name,
        normalize_descriptor(housing_type, gender),
        normalize_address(address),
    )


def parse_pdf_groups(pdf_path: Path) -> dict[tuple[str, str, str, str, str], dict]:
    def should_skip(entries: list[tuple[float, str]]) -> bool:
        return any("2024-2차" in text or "공급호수" in text or "구분 자치구" in text for _, text in entries)

    def parse_left_meta(left_text: str) -> tuple[str | None, str | None]:
        tokens = left_text.split()
        if len(tokens) >= 2 and tokens[0] in {"재공급", "신규공급"}:
            return tokens[0], tokens[1]
        if len(tokens) == 1:
            merged = re.match(r"^(재공급|신규공급)(.+)$", tokens[0])
            if merged:
                return merged.group(1), merged.group(2)
        return None, None

    raw_blocks: list[dict] = []
    current_block: list[list[tuple[float, str]]] = []

    reader = PdfReader(str(pdf_path))
    for page in reader.pages:
        for _, entries in group_page_fragments(page):
            if should_skip(entries):
                continue

            current_block.append(entries)
            rank_entry = next((parse_rank_line(text) for _, text in entries if parse_rank_line(text)), None)
            if not rank_entry or rank_entry[0] != "소계":
                continue

            block_info = {
                "counts": {"rank1Applicants": 0, "rank2Applicants": 0, "rank3Applicants": 0},
                "fullBase": None,
                "typeMeta": None,
            }

            for block_entries in current_block:
                rank_info = next((parse_rank_line(text) for _, text in block_entries if parse_rank_line(text)), None)
                address = next((text for x, text in block_entries if text.startswith("서울특별시")), None)
                left_text = " ".join(text for x, text in block_entries if x < 120)
                housing_name = next((text for x, text in block_entries if 440 <= x < 520), None)
                type_text = next((text for x, text in block_entries if 520 <= x < 575), None)
                gender_text = next((text for x, text in block_entries if 575 <= x < 620), None)
                supply_text = next((text for x, text in block_entries if 645 <= x < 680 and re.fullmatch(r"\d+", text)), None)

                if rank_info and rank_info[1] is not None:
                    rank_label, applicants = rank_info
                    if rank_label == "1순위":
                        block_info["counts"]["rank1Applicants"] = applicants
                    elif rank_label == "2순위":
                        block_info["counts"]["rank2Applicants"] = applicants
                    elif rank_label == "3순위":
                        block_info["counts"]["rank3Applicants"] = applicants

                if address and housing_name:
                    kind, district = parse_left_meta(left_text)
                    if kind and district:
                        housing_type, gender = split_type_gender(type_text, gender_text)
                        block_info["fullBase"] = {
                            "kind": kind,
                            "district": district,
                            "housingName": housing_name.strip(),
                            "address": address.strip(),
                            "housingType": housing_type,
                            "gender": gender,
                            "supply": int(supply_text) if supply_text else None,
                        }

                if rank_info is None and (type_text or gender_text or supply_text):
                    housing_type, gender = split_type_gender(type_text, gender_text)
                    block_info["typeMeta"] = {
                        "housingType": housing_type,
                        "gender": gender,
                        "supply": int(supply_text) if supply_text else None,
                    }

            raw_blocks.append(block_info)
            current_block = []

    groups: dict[tuple[str, str, str, str, str], dict] = {}
    next_full_bases: list[dict | None] = [None] * len(raw_blocks)
    next_base = None
    for index in range(len(raw_blocks) - 1, -1, -1):
        if raw_blocks[index]["fullBase"]:
            next_base = raw_blocks[index]["fullBase"]
        next_full_bases[index] = next_base

    previous_base = None
    for index, block in enumerate(raw_blocks):
        base = block["fullBase"] or next_full_bases[index] or previous_base
        type_meta = block["typeMeta"] or {}
        supply = type_meta.get("supply")
        housing_type = type_meta.get("housingType")
        gender = type_meta.get("gender")

        if supply is None and block["fullBase"] and block["fullBase"].get("supply") is not None:
            supply = block["fullBase"]["supply"]
            housing_type = housing_type if housing_type is not None else block["fullBase"].get("housingType")
            gender = gender if gender is not None else block["fullBase"].get("gender")

        if not base or supply is None:
            continue

        previous_base = block["fullBase"] or previous_base
        key = build_group_key(
            base["kind"],
            base["district"],
            base["housingName"],
            housing_type,
            gender,
            base["address"],
        )
        applicants = (
            block["counts"]["rank1Applicants"]
            + block["counts"]["rank2Applicants"]
            + block["counts"]["rank3Applicants"]
        )
        groups[key] = {
            "_order": index,
            "kind": base["kind"],
            "district": base["district"],
            "housingName": base["housingName"],
            "housingType": housing_type,
            "gender": gender,
            "address": base["address"],
            "supply": supply,
            "rank1Applicants": block["counts"]["rank1Applicants"],
            "rank2Applicants": block["counts"]["rank2Applicants"],
            "rank3Applicants": block["counts"]["rank3Applicants"],
            "applicants": applicants,
            "competitionRatio": round_half_up(applicants / supply, 1),
        }

    return groups


def merge_records(xlsx_records: list[dict], pdf_groups: dict[tuple[str, str, str, str, str], dict]) -> list[dict]:
    def base_key(record: dict) -> tuple[str, str, str, str]:
        return (
            record["kind"],
            record["district"],
            record["housingName"],
            normalize_address(record["address"]),
        )

    merged: list[dict] = []
    used_pdf_keys: set[tuple[str, str, str, str, str]] = set()
    pdf_by_base: dict[tuple[str, str, str, str], list[tuple[tuple[str, str, str, str, str], dict]]] = {}
    xlsx_by_base: dict[tuple[str, str, str, str], list[dict]] = {}

    for key, pdf_entry in pdf_groups.items():
        pdf_by_base.setdefault(base_key(pdf_entry), []).append((key, pdf_entry))
    for candidates in pdf_by_base.values():
        candidates.sort(key=lambda item: item[1].get("_order", 0))

    for record in xlsx_records:
        xlsx_by_base.setdefault(base_key(record), []).append(record)
    for records in xlsx_by_base.values():
        records.sort(key=lambda item: item.get("_order", 0))

    globally_unmatched_records: list[dict] = []

    for group_key, records in xlsx_by_base.items():
        candidates = pdf_by_base.get(group_key, [])
        remaining_records: list[dict] = []

        for record in records:
            exact_key = build_group_key(
                record["kind"],
                record["district"],
                record["housingName"],
                record.get("housingType"),
                record.get("gender"),
                record["address"],
            )
            pdf_entry = pdf_groups.get(exact_key)
            if pdf_entry and exact_key not in used_pdf_keys:
                used_pdf_keys.add(exact_key)
                merged.append(
                    {
                        **record,
                        "supply": pdf_entry["supply"],
                        "applicants": pdf_entry["applicants"],
                        "competitionRatio": pdf_entry["competitionRatio"],
                        "rank1Applicants": pdf_entry["rank1Applicants"],
                        "rank2Applicants": pdf_entry["rank2Applicants"],
                        "rank3Applicants": pdf_entry["rank3Applicants"],
                    }
                )
            else:
                remaining_records.append(record)

        remaining_candidates = [(key, entry) for key, entry in candidates if key not in used_pdf_keys]
        if remaining_records:
            if len(remaining_candidates) == len(remaining_records):
                for record, (candidate_key, pdf_entry) in zip(remaining_records, remaining_candidates):
                    used_pdf_keys.add(candidate_key)
                    merged.append(
                        {
                            **record,
                            "supply": pdf_entry["supply"],
                            "applicants": pdf_entry["applicants"],
                            "competitionRatio": pdf_entry["competitionRatio"],
                            "rank1Applicants": pdf_entry["rank1Applicants"],
                            "rank2Applicants": pdf_entry["rank2Applicants"],
                            "rank3Applicants": pdf_entry["rank3Applicants"],
                        }
                    )
            else:
                globally_unmatched_records.extend(remaining_records)

    remaining_pdf = [(key, entry) for key, entry in pdf_groups.items() if key not in used_pdf_keys]
    globally_unmatched_records.sort(key=lambda item: item.get("_order", 0))

    if globally_unmatched_records:
        for record in globally_unmatched_records:
            preferred = [
                item for item in remaining_pdf
                if item[1]["kind"] == record["kind"] and item[1]["district"] == record["district"]
            ]
            pool = preferred or remaining_pdf
            if not pool:
                raise ValueError(f"No pdf blocks left for fallback record: {record}")

            candidate_key, pdf_entry = min(
                pool,
                key=lambda item: abs(item[1].get("_order", 0) - record.get("_order", 0)),
            )
            remaining_pdf.remove((candidate_key, pdf_entry))
            used_pdf_keys.add(candidate_key)
            merged.append(
                {
                    **record,
                    "supply": pdf_entry["supply"],
                    "applicants": pdf_entry["applicants"],
                    "competitionRatio": pdf_entry["competitionRatio"],
                    "rank1Applicants": pdf_entry["rank1Applicants"],
                    "rank2Applicants": pdf_entry["rank2Applicants"],
                    "rank3Applicants": pdf_entry["rank3Applicants"],
                }
            )

    merged.sort(key=lambda record: record.get("_order", 0))
    return merged


def verify_records(records: list[dict]) -> None:
    keys = []
    for record in records:
        record_key = (
            record["roundKey"],
            record["kind"],
            record["district"],
            record["housingName"],
            record.get("housingType"),
            record.get("gender"),
            record["address"],
        )
        keys.append(record_key)

        total = record["rank1Applicants"] + record["rank2Applicants"] + record["rank3Applicants"]
        if total != record["applicants"]:
            raise ValueError(f"Priority count mismatch: {record_key}")

        expected_ratio = round_half_up(record["applicants"] / record["supply"], 1)
        if expected_ratio != round_half_up(record["competitionRatio"], 1):
            raise ValueError(f"Competition ratio mismatch: {record_key}")

    duplicates = [key for key, count in Counter(keys).items() if count > 1]
    if duplicates:
        raise ValueError(f"Duplicate records found: {duplicates[:5]}")


def merge_into_dataset(existing_records: list[dict], new_records: list[dict], round_key: str) -> list[dict]:
    filtered = [record for record in existing_records if record.get("roundKey") != round_key]
    combined = filtered + new_records
    combined.sort(
        key=lambda record: (
            str(record.get("noticeDate") or ""),
            record.get("district") or "",
            record.get("housingName") or "",
            record.get("housingType") or "",
        ),
        reverse=False,
    )
    combined.sort(key=lambda record: str(record.get("noticeDate") or ""), reverse=True)
    return combined


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import a market lab round from xlsx+pdf.")
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--pdf", type=Path, required=True)
    parser.add_argument("--round-key", required=True)
    parser.add_argument("--round-label", required=True)
    parser.add_argument("--notice-date", required=True)
    parser.add_argument("--check-only", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_records = parse_xlsx_records(args.xlsx, args.round_key, args.round_label, args.notice_date)
    pdf_groups = parse_pdf_groups(args.pdf)
    merged_records = merge_records(xlsx_records, pdf_groups)
    verify_records(merged_records)

    if not args.check_only:
        existing = load_records()
        save_records(merge_into_dataset(existing, merged_records, args.round_key))

    print(
        json.dumps(
            {
                "roundKey": args.round_key,
                "xlsxRecords": len(xlsx_records),
                "pdfGroups": len(pdf_groups),
                "mergedRecords": len(merged_records),
                "checkOnly": args.check_only,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
